import os
import json
from typing import Any

from openpyxl import Workbook, load_workbook

PATH = "config.json"
PATH_OUT = "config_out.json"
FILE_NAME = 'au.xlsx'


class Config:
    ignore: int
    sn: int
    threshold: int
    group: int
    name: str


class AU():
    _typeName: str
    _sut: str
    _sun: int
    config: Config
    _deleted: bool
    _subUnits: dict


def read_file_json(puth):
    with open(puth, 'r') as f:
        parsed = json.load(f)
    # print(json.dumps(parsed, indent=4))
    arr = parsed["_subUnits"]["_Box"]["11"]["_subUnits"]["Module"]["3"]["_subUnits"]["AL"]["1"]["_subUnits"]["AU"]
    # for i in arr:
    #     print(arr[i])
    # print(json.dumps(arr, indent=4))
    return arr


def set_header_sheets(sheet):
    sheet.cell(row=1, column=1, value="Addr")
    sheet.cell(row=1, column=2, value="Type")
    sheet.cell(row=1, column=3, value="_sut")
    sheet.cell(row=1, column=4, value="_sun")
    sheet.cell(row=1, column=5, value="ignore")
    sheet.cell(row=1, column=6, value="SN")
    sheet.cell(row=1, column=7, value="threshold")
    sheet.cell(row=1, column=8, value="group")
    sheet.cell(row=1, column=9, value="name")
    sheet.cell(row=1, column=10, value="_deleted")
    sheet.cell(row=1, column=11, value="_subUnits")


def write_file_xlsx(list_au):
    count = 2
    wb = Workbook()
    # ws = wb.active
    ws1 = wb.create_sheet("AL1",0)
    ws2 = wb.create_sheet("AL2", 1)
    set_header_sheets(ws1)

    for key, value in list_au.items():
        ws1.cell(row=count, column=1, value=key)
        ws1.cell(row=count, column=2, value=value["_typeName"])
        ws1.cell(row=count, column=3, value=value["_sut"])
        ws1.cell(row=count, column=4, value=value["_sun"])
        ws1.cell(row=count, column=5, value=value["config"]["ignore"])
        ws1.cell(row=count, column=6, value=value["config"]["SN"])
        ws1.cell(row=count, column=7, value=value["config"]["threshold"])
        ws1.cell(row=count, column=8, value=value["config"]["group"])
        ws1.cell(row=count, column=9, value=value["config"]["name"])
        ws1.cell(row=count, column=10, value=value["_deleted"])
        ws1.cell(row=count, column=11, value=None)
        count += 1
    wb.save('au.xlsx')


def read_file_xlsx():
    au_list = {}
    wb = load_workbook(filename=FILE_NAME)
    ws1 = wb["AL1"]
    for x in range(2,256):
        au = {}
        num_row = ws1.cell(row=x, column=1).value
        if num_row != None:
            au["_typeName"] = ws1.cell(row=x, column=2).value
            au["_sut"] = ws1.cell(row=x, column=3).value
            au["_sun"] = ws1.cell(row=x, column=1).value
            config = {}
            config["ignore"] = ws1.cell(row=x, column=5).value
            config["SN"] = ws1.cell(row=x, column=6).value
            config["threshold"] = ws1.cell(row=x, column=7).value
            config["group"] = ws1.cell(row=x, column=8).value
            name = ws1.cell(row=x, column=9).value
            if name != None:
                config["name"] = name
            else:
                config["name"] = ""
            au["config"] = config
            au["_deleted"] = ws1.cell(row=x, column=10).value
            if ws1.cell(row=x, column=11).value == None:
                au["_subUnits"] = {}
            else:
                au["_subUnits"] = {}
            au_list[str(num_row)] = au
    # print(au_list)
    return au_list


def read_json_file():
    with open(PATH, 'r') as f_in, open(PATH_OUT, "w") as f_out:
        list_au = read_file_xlsx()
        parsed = json.load(f_in)
        parsed["_subUnits"]["_Box"]["11"]["_subUnits"]["Module"]["3"]["_subUnits"]["AL"]["1"]["_subUnits"]["AU"] = list_au
        json.dump(parsed, f_out)

def main():
    if not os.path.exists(FILE_NAME):
        dump_au = read_file_json(PATH)
        write_file_xlsx(dump_au)
    read_json_file()


if __name__ == '__main__':
    main()
